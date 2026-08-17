# -*- coding: utf-8 -*-
import openpyxl
p = r"D:/projects/voc_platform/data/exports/重打500_20260817_125739.xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
print("sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"--- {sn}: max_row={ws.max_row}, max_col={ws.max_column}")
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(min_row=1, max_row=6, values_only=True)
for i, r in enumerate(rows):
    print(f"ROW{i+1}:", r)
wb.close()
