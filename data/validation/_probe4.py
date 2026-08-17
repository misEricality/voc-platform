# -*- coding: utf-8 -*-
import openpyxl
from collections import Counter
p = r"D:/projects/voc_platform/data/exports/重打500_20260817_125739.xlsx"
wb = openpyxl.load_workbook(p, read_only=True, data_only=True)

# opinions
ws = wb["opinions"]
op_rows = list(ws.iter_rows(min_row=2, values_only=True))
print("opinions rows:", len(op_rows))
sent = Counter(r[6] for r in op_rows)
print("opinion_sentiment:", dict(sent))
l1 = Counter((r[8] or "").split("/")[0] for r in op_rows)
print("\nfull_path L1:", dict(l1))
l2 = Counter("/".join((r[8] or "").split("/")[:2]) for r in op_rows)
print("\nfull_path L2:")
for k, v in l2.most_common(30):
    print(f"  {v:4d}  {k}")
# 有没有 phrase 为空的
empty = sum(1 for r in op_rows if not (r[9] and str(r[9]).strip()))
print("\nopinions empty phrase:", empty)

# comments
ws2 = wb["comments"]
c_rows = list(ws2.iter_rows(min_row=2, values_only=True))
print("\ncomments rows:", len(c_rows))
c_sent = Counter(r[5] for r in c_rows)
print("comment sentiment:", dict(c_sent))
c_topic = Counter(r[8] for r in c_rows)
print("comment topic:", dict(c_topic))
c_sub = Counter(r[9] for r in c_rows if r[9])
print("comment sub_topics 有值行数:", sum(1 for r in c_rows if r[9]))
wb.close()
