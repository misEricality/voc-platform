"""按固定评论 ID 列表导出重打结果为 xlsx。

依赖 workspace 自带 Python（含 openpyxl），项目 .venv-review 未装 openpyxl。
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


STEAM_APP_NAMES = {
    "730": "Counter-Strike 2",
    "570": "Dota 2",
    "578080": "PUBG: BATTLEGROUNDS",
    "1172470": "Apex Legends",
    "2358720": "黑神话：悟空",
    "1222140": "底特律：化身为人",
    "292030": "巫师 3：狂猎",
    "289070": "文明 6",
    "1903340": "光与影：33号远征队",
    "753640": "星际拓荒",
}


def resolve_target_name(platform: str, target_id: str, meta_name: str | None) -> str:
    if meta_name:
        return meta_name
    if platform == "steam":
        appid = target_id.removeprefix("steam:")
        return STEAM_APP_NAMES.get(appid, target_id)
    return target_id or ""


def load_ids(path: Path) -> list[int]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [int(x) for x in data]


def fetch(conn: sqlite3.Connection, ids: list[int]):
    ph = ",".join("?" for _ in ids)
    comments = conn.execute(
        f"""
        SELECT c.id, c.platform, c.source_id, c.target_id,
               json_extract(c.extra_meta, '$.name') AS target_name,
               c.content, c.sentiment, c.sentiment_score, c.sentiment_confidence,
               c.topic, c.sub_topics, c.analyzed_at
        FROM comments c
        WHERE c.id IN ({ph})
        ORDER BY c.id
        """,
        ids,
    ).fetchall()
    opinions = conn.execute(
        f"""
        SELECT o.id, o.comment_id, o.full_path, o.sentiment, o.sentiment_confidence,
               o.quote, o.quote_start, o.quote_end
        FROM comment_opinions o
        WHERE o.comment_id IN ({ph})
        ORDER BY o.comment_id, o.id
        """,
        ids,
    ).fetchall()
    return comments, opinions


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    widths = {
        "A": 12, "B": 14, "C": 18, "D": 18, "E": 18, "F": 12, "G": 10,
        "H": 12, "I": 16, "J": 30, "K": 12, "L": 14, "M": 80,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ids", default="data/validation/_random500_ids.json")
    parser.add_argument("--db", default="data/voc.db")
    parser.add_argument("--out", default=None)
    args = parser.parse_args()

    ids = load_ids(Path(args.ids))
    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row
    comments, opinions = fetch(conn, ids)

    opinion_counts = {}
    for op in opinions:
        opinion_counts[op["comment_id"]] = opinion_counts.get(op["comment_id"], 0) + 1
    comments_map = {c["id"]: c for c in comments}

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "comments"
    ws1.append([
        "comment_id", "platform", "source_id", "target_id", "target_name",
        "sentiment", "sentiment_score", "sentiment_confidence", "topic",
        "opinion_count", "content",
    ])
    for c in comments:
        target_name = resolve_target_name(
            c["platform"], c["target_id"], c["target_name"]
        )
        ws1.append([
            c["id"], c["platform"], c["source_id"], c["target_id"],
            target_name, c["sentiment"] or "",
            round(c["sentiment_score"] or 0, 2),
            round(c["sentiment_confidence"] or 0, 2),
            c["topic"] or "",
            opinion_counts.get(c["id"], 0), c["content"] or "",
        ])
    style_header(ws1)

    ws2 = wb.create_sheet("opinions")
    ws2.append([
        "opinion_id", "comment_id", "source_id", "target_id", "target_name",
        "comment_sentiment", "opinion_sentiment", "opinion_confidence",
        "full_path", "phrase", "quote_start", "quote_end",
    ])
    for op in opinions:
        c = comments_map.get(op["comment_id"])
        if not c:
            continue
        target_name = resolve_target_name(
            c["platform"], c["target_id"], c["target_name"]
        )
        ws2.append([
            op["id"], op["comment_id"], c["source_id"], c["target_id"],
            target_name, c["sentiment"] or "", op["sentiment"],
            round(op["sentiment_confidence"] or 0, 2)
            if op["sentiment_confidence"] is not None
            else "",
            op["full_path"], op["quote"], op["quote_start"], op["quote_end"],
        ])
    style_header(ws2)

    if args.out:
        out = Path(args.out)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = Path(f"data/exports/重打500_{ts}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ 已导出 {len(comments)} 条评论 + {len(opinions)} 个观点")
    print(f"📁 {out}")
    conn.close()


if __name__ == "__main__":
    main()
