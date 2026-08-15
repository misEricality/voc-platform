"""导出 B 站采集结果（评论 + 弹幕）为 .xlsx

用法：
    python scripts/dev/export_bilibili.py [--out 路径]
"""
from __future__ import annotations

import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def get_conn(db: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db)
    conn.row_factory = sqlite3.Row
    return conn


def style_header(ws, ncols: int) -> None:
    """表头样式：深色底 + 白字"""
    fill = PatternFill("solid", fgColor="2F5496")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_comments_sheet(ws, conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        """
        SELECT c.id, c.source_id, c.target_id, c.author, c.author_id,
               c.likes, c.replies, c.posted_at, c.content, c.extra_json
        FROM comments c
        WHERE c.platform = 'bilibili'
        ORDER BY c.likes DESC
        """
    ).fetchall()
    ws.append([
        "comment_id", "source_id(rpid)", "target_id", "作者昵称", "mid",
        "点赞数", "楼中楼", "评论时间", "Lv", "性別", "VIP", "认证", "评论内容",
    ])
    for r in rows:
        prof = {}
        if r["extra_json"]:
            try:
                prof = (json.loads(r["extra_json"]) or {}).get("profile") or {}
            except Exception:
                pass
        vip = prof.get("vip") or {}
        official = prof.get("official") or {}
        ws.append([
            r["id"], r["source_id"], r["target_id"], r["author"], r["author_id"],
            r["likes"], r["replies"], r["posted_at"],
            prof.get("level"), prof.get("sex"),
            vip.get("status", 0), official.get("title") or "", r["content"],
        ])
    return len(rows)


def build_danmaku_sheet(ws, conn: sqlite3.Connection) -> int:
    rows = conn.execute(
        """
        SELECT id, video_id, cid, progress, mode, color, user_hash, posted_at, content
        FROM danmaku
        ORDER BY progress
        """
    ).fetchall()
    ws.append(["id", "video_id", "cid", "视频内时间(秒)", "类型", "颜色", "用户hash", "发送时间", "弹幕内容"])
    for r in rows:
        ws.append([
            r["id"], r["video_id"], r["cid"], r["progress"], r["mode"],
            r["color"], r["user_hash"], r["posted_at"], r["content"],
        ])
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="导出 B 站采集结果")
    parser.add_argument("--db", default="data/voc.db")
    parser.add_argument("--out", default=None, help="输出路径（默认 data/exports/bilibili_时间戳.xlsx）")
    args = parser.parse_args()

    conn = get_conn(args.db)
    out = args.out or f"data/exports/bilibili_{Path(args.db).stem}.xlsx"
    Path(out).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "comments"
    n1 = build_comments_sheet(ws1, conn)
    style_header(ws1, ws1.max_column)
    # 列宽
    widths1 = [10, 14, 26, 12, 12, 8, 8, 20, 6, 6, 6, 14, 60]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("danmaku")
    n2 = build_danmaku_sheet(ws2, conn)
    style_header(ws2, ws2.max_column)
    widths2 = [8, 26, 12, 12, 6, 8, 10, 20, 60]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(out)
    conn.close()
    print(f"✅ 已导出: {out}")
    print(f"  Sheet1 comments: {n1} 条（按点赞降序）")
    print(f"  Sheet2 danmaku:  {n2} 条（按视频内时间升序）")


if __name__ == "__main__":
    main()
