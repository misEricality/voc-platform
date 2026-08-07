"""导出标注结果为 .xlsx（2026-08-05 一次性任务）

生成两个 Sheet：

**Sheet 1: comments（每条评论一行）**
- comment_id（DB 主键）
- source_id（Steam recommendationid，原声 id）
- target_id
- target_name
- sentiment
- topic（L1）
- sub_topics（L2 列表）
- opinion_count
- content（原声）

**Sheet 2: opinions（每个观点一行）**
- comment_id（外键）
- source_id
- target_id
- target_name
- sentiment
- label_level（L1/L2）
- label
- full_path（L1/L2 拼接）
- quote（观点）
- quote_start
- quote_end
- content（原声，简短版）

默认输出到 `data/exports/voc_export_YYYYMMDD_HHMMSS.xlsx`
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def get_conn(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def fetch_all(
    conn: sqlite3.Connection, only_with_opinions: bool = False
) -> tuple[list[sqlite3.Row], list[sqlite3.Row]]:
    """查评论 + opinions

    Args:
        only_with_opinions: True 时只查有 opinions 的评论（抽样验证场景）
    """
    if only_with_opinions:
        comments = list(conn.execute("""
            SELECT c.id, c.platform, c.source_id, c.target_id,
                   json_extract(c.extra_meta, '$.name') as target_name,
                   c.content, c.sentiment, c.sentiment_score, c.sentiment_confidence,
                   c.topic, c.sub_topics, c.analyzed_at
            FROM comments c
            WHERE c.analyzed_at IS NOT NULL
              AND EXISTS (
                  SELECT 1 FROM comment_opinions o WHERE o.comment_id = c.id
              )
            ORDER BY c.id
        """))
    else:
        comments = list(conn.execute("""
            SELECT c.id, c.platform, c.source_id, c.target_id,
                   json_extract(c.extra_meta, '$.name') as target_name,
                   c.content, c.sentiment, c.sentiment_score, c.sentiment_confidence,
                   c.topic, c.sub_topics, c.analyzed_at
            FROM comments c
            WHERE c.analyzed_at IS NOT NULL
            ORDER BY c.id
        """))
    opinions = list(conn.execute("""
        SELECT id, comment_id, full_path, sentiment, quote, quote_start, quote_end
        FROM comment_opinions
        ORDER BY comment_id, id
    """))
    return comments, opinions


def build_comments_rows(
    comments: list[sqlite3.Row],
    opinion_counts: dict[int, int],
) -> list[list]:
    """构造 Sheet 1 行"""
    rows = []
    for c in comments:
        try:
            subs = json.loads(c["sub_topics"]) if c["sub_topics"] else []
        except Exception:
            subs = []
        rows.append([
            c["id"],
            c["source_id"],
            c["target_id"],
            c["target_name"] or "",
            c["sentiment"] or "",
            round(c["sentiment_score"] or 0, 2),
            round(c["sentiment_confidence"] or 0, 2),
            c["topic"] or "",
            "、".join(subs),
            opinion_counts.get(c["id"], 0),
            c["content"] or "",
        ])
    return rows


def build_opinions_rows(
    comments_map: dict[int, sqlite3.Row],
    opinions: list[sqlite3.Row],
) -> list[list]:
    """构造 Sheet 2 行（v2：完整路径 + 观点级情感）"""
    rows = []
    for op in opinions:
        c = comments_map.get(op["comment_id"])
        if not c:
            continue
        # 截短 content（避免单元格过长）
        content_short = (c["content"] or "")[:120]
        if (c["content"] or "") and len(c["content"]) > 120:
            content_short += "..."
        rows.append([
            op["id"],
            op["comment_id"],
            c["source_id"],
            c["target_id"],
            c["target_name"] or "",
            c["sentiment"] or "",
            op["sentiment"],          # 观点级情感
            op["full_path"],          # 完整路径 L1/L2/L3
            op["quote"],
            op["quote_start"],
            op["quote_end"],
            content_short,
        ])
    return rows


def style_header(ws, row_count: int) -> None:
    """表头加粗 + 冻结首行 + 列宽自适应"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # 列宽：自动设一个合理初值
    widths = {
        "A": 12, "B": 16, "C": 14, "D": 16, "E": 12, "F": 10, "G": 12,
        "H": 16, "I": 30, "J": 12, "K": 12, "L": 14, "M": 80, "N": 80,
    }
    for col, w in widths.items():
        if col in ws.column_dimensions or ws.max_column >= ord(col) - ord("A") + 1:
            ws.column_dimensions[col].width = w


def write_xlsx(out_path: Path, comments_rows: list, opinions_rows: list) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: comments
    ws1 = wb.active
    ws1.title = "comments"
    ws1.append([
        "comment_id", "source_id", "target_id", "target_name",
        "sentiment", "sentiment_score", "sentiment_confidence",
        "topic", "sub_topics", "opinion_count", "content",
    ])
    for row in comments_rows:
        ws1.append(row)
    style_header(ws1, len(comments_rows))

    # Sheet 2: opinions（方案4：完整路径 + 观点情感 + phrase）
    ws2 = wb.create_sheet("opinions")
    ws2.append([
        "opinion_id", "comment_id", "source_id", "target_id", "target_name",
        "comment_sentiment", "opinion_sentiment", "full_path",
        "phrase", "quote_start", "quote_end", "content",
    ])
    for row in opinions_rows:
        ws2.append(row)
    style_header(ws2, len(opinions_rows))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="导出 VoC 标注结果为 .xlsx")
    parser.add_argument("--db", default="data/voc.db", help="数据库路径")
    parser.add_argument("--out", help="输出路径（默认 data/exports/voc_export_时间戳.xlsx）")
    parser.add_argument("--only-with-opinions", action="store_true", help="只导出有观点的评论（抽样验证）")
    parser.add_argument("--open", action="store_true", help="导出后用默认程序打开")
    args = parser.parse_args()

    conn = get_conn(args.db)
    comments, opinions = fetch_all(conn, only_with_opinions=args.only_with_opinions)

    # 统计 opinion 数量
    opinion_counts: dict[int, int] = {}
    for op in opinions:
        opinion_counts[op["comment_id"]] = opinion_counts.get(op["comment_id"], 0) + 1

    comments_map = {c["id"]: c for c in comments}

    comments_rows = build_comments_rows(comments, opinion_counts)
    opinions_rows = build_opinions_rows(comments_map, opinions)

    # 输出路径
    if args.out:
        out_path = Path(args.out)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = Path(f"data/exports/voc_export_{ts}.xlsx")

    write_xlsx(out_path, comments_rows, opinions_rows)
    print(f"✅ 已导出 {len(comments)} 条评论 + {len(opinions)} 个观点")
    print(f"📁 文件: {out_path}")
    print(f"   Sheet 1 'comments': {len(comments)} 行")
    print(f"   Sheet 2 'opinions': {len(opinions)} 行")

    if args.open:
        import os
        os.startfile(out_path)  # Windows-only
        print("🖱️ 已用默认程序打开")

    conn.close()


if __name__ == "__main__":
    main()